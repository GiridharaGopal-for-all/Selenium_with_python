import openpyxl

class xl:
    @staticmethod
    def excel_data():
        data_list = []
        file = openpyxl.load_workbook("E:\\Giri\\testing_excel.xlsx")
        sheet = file["Sheet1"]
        max_row = sheet.max_row
        for i in range(2, max_row + 1):
            username = sheet.cell(row=i, column=1).value
            password = sheet.cell(row=i, column=2).value
            data_list.append((username, password))
        return data_list